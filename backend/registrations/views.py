import io
import json
import re
from datetime import datetime

from django.db.models import Q
from django.http import HttpResponse, JsonResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods

from .forms import EventRegistrationForm
from .models import EventRegistration


XLSX_EXPORT_CONFIG = json.loads(
	"""
	{
	  "sheetName": "Registrations",
	  "header": {
		"height": 26,
		"font": {"name": "Calibri", "size": 11, "bold": true, "color": "FFFFFF"},
		"alignment": {"horizontal": "left", "vertical": "center", "wrapText": true, "indent": 0}
	  },
	  "rows": {
		"height": 22,
		"font": {"name": "Calibri", "size": 10, "bold": false, "color": "1F4736"},
		"alignment": {"horizontal": "left", "vertical": "top", "wrapText": true, "indent": 1}
	  },
	  "columns": [
		{"header": "Reference", "key": "reference", "width": 22},
		{"header": "Email", "key": "email", "width": 30},
		{"header": "Last Name", "key": "last_name", "width": 18},
		{"header": "First Name", "key": "first_name", "width": 18},
		{"header": "Middle Initial", "key": "middle_initial", "width": 14},
		{"header": "Designation", "key": "designation", "width": 24},
		{"header": "Mobile/CP No.", "key": "mobile_cp_no", "width": 16},
		{"header": "Viber No.", "key": "viber_no", "width": 16},
		{"header": "G-Cash No.", "key": "gcash_no", "width": 16},
		{"header": "Personal Email", "key": "personal_email_address", "width": 30},
		{"header": "LinkedIn Account", "key": "linkedin_account", "width": 28},
		{"header": "Facebook Account", "key": "facebook_account", "width": 28},
		{"header": "Messenger Account", "key": "messenger_account", "width": 28},
		{"header": "Company Name", "key": "company_name", "width": 28},
		{"header": "Company Category", "key": "company_category", "width": 18},
		{"header": "Industry Type", "key": "industry_type", "width": 22},
		{"header": "Company Office Address", "key": "company_office_address", "width": 34},
		{"header": "Company Landline No.", "key": "company_landline_no", "width": 18},
		{"header": "Company Email", "key": "company_email_address", "width": 30},
		{"header": "Submitted At", "key": "submitted_at", "width": 22}
	  ],
	  "themes": {
		"mint": {
		  "headerFill": "3F8657",
		  "oddRowFill": "F6FBF8",
		  "evenRowFill": "ECF5EF",
		  "border": "BFD3C5"
		},
		"gold": {
		  "headerFill": "B9923D",
		  "oddRowFill": "FFF8EC",
		  "evenRowFill": "FFF3DE",
		  "border": "E4D1A4"
		},
		"blue": {
		  "headerFill": "3E7BA6",
		  "oddRowFill": "F2F7FC",
		  "evenRowFill": "EAF2F9",
		  "border": "BFD2E0"
		}
	  }
	}
	"""
)

DUPLICATE_STATUS_RED = 'red'
DUPLICATE_STATUS_ORANGE = 'orange'
DUPLICATE_STATUS_YELLOW = 'yellow'
DUPLICATE_STATUS_GREEN = 'green'

DUPLICATE_STATUS_LABELS = {
	DUPLICATE_STATUS_RED: 'Red - Same GCash and Email/Personal Email',
	DUPLICATE_STATUS_ORANGE: 'Orange - Same Mobile/CP No.',
	DUPLICATE_STATUS_YELLOW: 'Yellow - Same Full Name',
	DUPLICATE_STATUS_GREEN: 'Green - Unique entry',
}

DUPLICATE_STATUS_FILLS = {
	DUPLICATE_STATUS_RED: 'FFF1F1',
	DUPLICATE_STATUS_ORANGE: 'FFF3E6',
	DUPLICATE_STATUS_YELLOW: 'FFFBE0',
	DUPLICATE_STATUS_GREEN: 'EAF8EE',
}


def _normalize_text(value: str) -> str:
	if not value:
		return ''
	return re.sub(r'\s+', ' ', str(value).strip().lower())


def _normalize_email(value: str) -> str:
	email = _normalize_text(value)
	if '@' not in email:
		return email

	local, domain = email.split('@', 1)
	is_gmail_domain = domain in {'gmail.com', 'googlemail.com'}
	if '+' in local and is_gmail_domain:
		local = local.split('+', 1)[0]
	if is_gmail_domain:
		local = local.replace('.', '')
	return f'{local}@{domain}'


def _normalize_name(value: str) -> str:
	name = _normalize_text(value)
	return re.sub(r'[^a-z\s]', '', name)


def _normalize_phone(value: str) -> str:
	if not value:
		return ''
	digits = ''.join(ch for ch in str(value) if ch.isdigit())
	if digits.startswith('63') and len(digits) == 12:
		return f"0{digits[2:]}"
	if len(digits) == 10 and not digits.startswith('0'):
		return f"0{digits}"
	return digits


def _build_registration_fingerprint(registration: EventRegistration):
	full_name = _normalize_name(
		f"{registration.first_name} {registration.middle_initial} {registration.last_name}"
	)

	identity_emails = {
		_normalize_email(registration.email),
		_normalize_email(registration.personal_email_address),
	}
	identity_emails.discard('')

	mobile_number = _normalize_phone(registration.mobile_cp_no)
	gcash_number = _normalize_phone(registration.gcash_no)

	return {
		'fullName': full_name,
		'identityEmails': sorted(identity_emails),
		'mobileNo': mobile_number,
		'gcashNo': gcash_number,
	}


def _evaluate_duplicate_signal(registration_id: int, fingerprints) -> tuple[int, str]:
	current = fingerprints.get(registration_id)
	if not current:
		return 0, DUPLICATE_STATUS_GREEN

	has_orange_match = False
	has_yellow_match = False

	for other_id, other in fingerprints.items():
		if other_id == registration_id:
			continue

		email_overlap = bool(set(current['identityEmails']).intersection(other['identityEmails']))
		same_gcash = bool(current['gcashNo']) and current['gcashNo'] == other['gcashNo']
		if same_gcash and email_overlap:
			return 100, DUPLICATE_STATUS_RED

		same_mobile = bool(current['mobileNo']) and current['mobileNo'] == other['mobileNo']
		if same_mobile:
			has_orange_match = True

		same_full_name = bool(current['fullName']) and current['fullName'] == other['fullName']
		if same_full_name:
			has_yellow_match = True

	if has_orange_match:
		return 80, DUPLICATE_STATUS_ORANGE
	if has_yellow_match:
		return 60, DUPLICATE_STATUS_YELLOW
	return 0, DUPLICATE_STATUS_GREEN


def _recompute_duplicate_metadata():
	registrations = list(EventRegistration.objects.all())
	if not registrations:
		return

	fingerprints = {
		registration.id: _build_registration_fingerprint(registration)
		for registration in registrations
	}

	changed = []
	for registration in registrations:
		score_value, status_value = _evaluate_duplicate_signal(registration.id, fingerprints)
		if registration.duplicate_score != score_value or registration.duplicate_status != status_value:
			registration.duplicate_score = score_value
			registration.duplicate_status = status_value
			changed.append(registration)

	if changed:
		EventRegistration.objects.bulk_update(changed, ['duplicate_score', 'duplicate_status'])


def _with_cors_headers(response: JsonResponse) -> JsonResponse:
	response['Access-Control-Allow-Origin'] = '*'
	response['Access-Control-Allow-Methods'] = 'GET, POST, OPTIONS'
	response['Access-Control-Allow-Headers'] = 'Content-Type'
	return response


def _format_reference(registration: EventRegistration) -> str:
	date_part = registration.created_at.strftime('%Y-%m%d')
	sequence = f'{registration.id:04d}'
	return f'MISI-{date_part}-{sequence}'


def _serialize_registration(registration: EventRegistration):
	full_name = ' '.join(part for part in [registration.first_name, registration.last_name] if part and part != '-').strip()
	return {
		'id': registration.id,
		'reference': _format_reference(registration),
		'refNumber': _format_reference(registration),
		'name': full_name,
		'privacyAccepted': registration.data_privacy_consent,
		'email': registration.email,
		'lastName': registration.last_name,
		'firstName': registration.first_name,
		'middleInitial': registration.middle_initial,
		'designation': registration.designation,
		'mobileNo': registration.mobile_cp_no,
		'viberNo': registration.viber_no,
		'gcashNo': registration.gcash_no,
		'personalEmail': registration.personal_email_address,
		'linkedinAccount': registration.linkedin_account,
		'facebookAccount': registration.facebook_account,
		'messengerAccount': registration.messenger_account,
		'companyName': registration.company_name,
		'companyCategory': registration.company_category,
		'industryType': registration.industry_type,
		'companyAddress': registration.company_office_address,
		'companyLandline': registration.company_landline_no,
		'companyEmail': registration.company_email_address,
		'companyIdToBring': registration.company_id_to_bring,
		'bringCompanyId': registration.company_id_to_bring,
		'gender': registration.company_category,
		'schoolUniversity': registration.company_name,
		'course': registration.industry_type,
		'yearLevel': registration.designation,
		'vehicleType': registration.vehicle_type,
		'willCome': registration.will_come,
		'attendeeCount': registration.attendee_count,
		'attendeeDetails': registration.additional_attendees,
		'duplicateScore': registration.duplicate_score,
		'duplicateStatus': registration.duplicate_status,
		'createdAt': registration.created_at.isoformat(),
		'date': registration.created_at.isoformat(),
	}


def _parse_bool_flag(value):
	if value is None:
		return None
	lower = value.strip().lower()
	if lower in {'true', '1', 'yes'}:
		return True
	if lower in {'false', '0', 'no'}:
		return False
	return None


def _map_payload(payload):
	full_name = (payload.get('name', '') or '').strip()
	name_parts = full_name.split()
	first_name = (payload.get('firstName', '') or '').strip() or (name_parts[0] if name_parts else '')
	last_name = (payload.get('lastName', '') or '').strip() or (' '.join(name_parts[1:]) if len(name_parts) > 1 else '-')
	middle_initial = (payload.get('middleInitial', '') or '').strip()
	mobile_number = payload.get('mobileNo', '')
	email = payload.get('email', '')
	gender = (payload.get('gender', '') or '').strip()
	school_university = (payload.get('schoolUniversity', '') or '').strip()
	course = (payload.get('course', '') or '').strip()
	year_level = (payload.get('yearLevel', '') or '').strip()

	return {
		'data_privacy_consent': payload.get('privacyAccepted'),
		'email': email,
		'last_name': last_name,
		'first_name': first_name,
		'middle_initial': middle_initial,
		'designation': year_level,
		'mobile_cp_no': mobile_number,
		'viber_no': payload.get('viberNo', ''),
		'gcash_no': payload.get('gcashNo') or mobile_number,
		'personal_email_address': payload.get('personalEmail') or email,
		'linkedin_account': payload.get('linkedinAccount', ''),
		'facebook_account': payload.get('facebookAccount', ''),
		'messenger_account': payload.get('messengerAccount', ''),
		'company_name': payload.get('companyName') or school_university,
		'company_category': payload.get('companyCategory') or gender,
		'industry_type': payload.get('industryType') or course,
		'company_office_address': payload.get('companyAddress') or 'N/A',
		'company_landline_no': payload.get('companyLandline') or 'N/A',
		'company_email_address': payload.get('companyEmail') or email,
		'company_id_to_bring': payload.get('companyIdToBring'),
		'vehicle_type': payload.get('vehicleType', ''),
		'will_come': True,
		'attendee_count': 1,
		'additional_attendees': [],
	}


def _build_export_row(registration: EventRegistration):
	return {
		'reference': _format_reference(registration),
		'email': registration.email,
		'last_name': registration.last_name,
		'first_name': registration.first_name,
		'middle_initial': registration.middle_initial,
		'designation': registration.designation,
		'mobile_cp_no': registration.mobile_cp_no,
		'viber_no': registration.viber_no,
		'gcash_no': registration.gcash_no,
		'personal_email_address': registration.personal_email_address,
		'linkedin_account': registration.linkedin_account,
		'facebook_account': registration.facebook_account,
		'messenger_account': registration.messenger_account,
		'company_name': registration.company_name,
		'company_category': registration.company_category,
		'industry_type': registration.industry_type,
		'company_office_address': registration.company_office_address,
		'company_landline_no': registration.company_landline_no,
		'company_email_address': registration.company_email_address,
		'company_id_to_bring': 'Yes' if registration.company_id_to_bring else 'No',
		'vehicle_type': registration.vehicle_type,
		'will_come': 'Yes' if registration.will_come else 'No',
		'attendee_count': registration.attendee_count,
		'additional_attendees': json.dumps(registration.additional_attendees, ensure_ascii=True),
		'duplicate_status': registration.duplicate_status,
		'submitted_at': registration.created_at.strftime('%Y-%m-%d %H:%M:%S'),
	}


@csrf_exempt
@require_http_methods(['POST', 'OPTIONS'])
def register_api(request):
	if request.method == 'OPTIONS':
		return _with_cors_headers(JsonResponse({'ok': True}))

	try:
		payload = json.loads(request.body.decode('utf-8'))
	except (json.JSONDecodeError, UnicodeDecodeError):
		return _with_cors_headers(JsonResponse({'error': 'Invalid JSON payload.'}, status=400))

	form = EventRegistrationForm(_map_payload(payload))
	if not form.is_valid():
		return _with_cors_headers(JsonResponse({'success': False, 'errors': form.errors}, status=400))

	registration = form.save()
	_recompute_duplicate_metadata()
	registration.refresh_from_db(fields=['duplicate_score', 'duplicate_status'])
	response_payload = {
		'success': True,
		'reference': _format_reference(registration),
		'createdAt': registration.created_at.isoformat(),
		'firstName': registration.first_name,
		'lastName': registration.last_name,
		'email': registration.email,
	}
	return _with_cors_headers(JsonResponse(response_payload, status=201))


@require_http_methods(['GET'])
def manage_registrations_api(request):
	_recompute_duplicate_metadata()
	registrations = EventRegistration.objects.all()

	search_query = request.GET.get('q', '').strip()
	if search_query:
		registrations = registrations.filter(
			Q(first_name__icontains=search_query)
			| Q(last_name__icontains=search_query)
			| Q(email__icontains=search_query)
			| Q(company_name__icontains=search_query)
			| Q(vehicle_type__icontains=search_query)
		)

	will_come_flag = _parse_bool_flag(request.GET.get('will_come'))
	if will_come_flag is not None:
		registrations = registrations.filter(will_come=will_come_flag)

	date_from = request.GET.get('date_from', '').strip()
	if date_from:
		try:
			parsed_date_from = datetime.strptime(date_from, '%Y-%m-%d').date()
			registrations = registrations.filter(created_at__date__gte=parsed_date_from)
		except ValueError:
			pass

	date_to = request.GET.get('date_to', '').strip()
	if date_to:
		try:
			parsed_date_to = datetime.strptime(date_to, '%Y-%m-%d').date()
			registrations = registrations.filter(created_at__date__lte=parsed_date_to)
		except ValueError:
			pass

	try:
		page = int(request.GET.get('page', '1'))
	except ValueError:
		page = 1
	page = max(page, 1)

	try:
		page_size = int(request.GET.get('page_size', '10'))
	except ValueError:
		page_size = 10
	page_size = min(max(page_size, 1), 100)

	total = registrations.count()
	total_pages = max((total + page_size - 1) // page_size, 1)
	if page > total_pages:
		page = total_pages

	start_index = (page - 1) * page_size
	end_index = start_index + page_size
	paged_results = registrations[start_index:end_index]

	results = [_serialize_registration(item) for item in paged_results]
	return JsonResponse(
		{
			'results': results,
			'pagination': {
				'page': page,
				'pageSize': page_size,
				'total': total,
				'totalPages': total_pages,
				'hasNext': page < total_pages,
				'hasPrevious': page > 1,
			},
		}
	)


@require_http_methods(['GET'])
def export_registrations_xlsx(request):
	_recompute_duplicate_metadata()
	theme_key = request.GET.get('theme', 'mint').strip().lower()
	themes = XLSX_EXPORT_CONFIG.get('themes', {})
	theme = themes.get(theme_key, themes.get('mint', {}))
	columns = XLSX_EXPORT_CONFIG.get('columns', [])
	header_config = XLSX_EXPORT_CONFIG.get('header', {})
	row_config = XLSX_EXPORT_CONFIG.get('rows', {})

	header_fill = PatternFill(fill_type='solid', fgColor=f"FF{theme.get('headerFill', '3F8657')}")
	border_side = Side(style='thin', color=f"FF{theme.get('border', 'BFD3C5')}")
	cell_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

	header_font = Font(
		name=header_config.get('font', {}).get('name', 'Calibri'),
		size=header_config.get('font', {}).get('size', 11),
		bold=header_config.get('font', {}).get('bold', True),
		color=f"FF{header_config.get('font', {}).get('color', 'FFFFFF')}",
	)
	data_font = Font(
		name=row_config.get('font', {}).get('name', 'Calibri'),
		size=row_config.get('font', {}).get('size', 10),
		bold=row_config.get('font', {}).get('bold', False),
		color=f"FF{row_config.get('font', {}).get('color', '1F4736')}",
	)

	header_alignment = Alignment(
		horizontal=header_config.get('alignment', {}).get('horizontal', 'left'),
		vertical=header_config.get('alignment', {}).get('vertical', 'center'),
		wrap_text=header_config.get('alignment', {}).get('wrapText', True),
		indent=header_config.get('alignment', {}).get('indent', 0),
	)
	data_alignment = Alignment(
		horizontal=row_config.get('alignment', {}).get('horizontal', 'left'),
		vertical=row_config.get('alignment', {}).get('vertical', 'top'),
		wrap_text=row_config.get('alignment', {}).get('wrapText', True),
		indent=row_config.get('alignment', {}).get('indent', 1),
	)

	workbook = Workbook()
	sheet = workbook.active
	sheet.title = XLSX_EXPORT_CONFIG.get('sheetName', 'Registrations')
	sheet.freeze_panes = 'A2'
	sheet.row_dimensions[1].height = header_config.get('height', 26)

	for index, column in enumerate(columns, start=1):
		cell = sheet.cell(row=1, column=index, value=column.get('header', ''))
		cell.fill = header_fill
		cell.font = header_font
		cell.alignment = header_alignment
		cell.border = cell_border
		sheet.column_dimensions[get_column_letter(index)].width = column.get('width', 18)

	for row_index, registration in enumerate(EventRegistration.objects.all(), start=2):
		sheet.row_dimensions[row_index].height = row_config.get('height', 22)
		export_row = _build_export_row(registration)
		status_fill_hex = DUPLICATE_STATUS_FILLS.get(export_row.get('duplicate_status', ''), DUPLICATE_STATUS_FILLS[DUPLICATE_STATUS_GREEN])
		row_fill = PatternFill(fill_type='solid', fgColor=f'FF{status_fill_hex}')

		for column_index, column in enumerate(columns, start=1):
			value = export_row.get(column.get('key', ''), '')
			cell = sheet.cell(row=row_index, column=column_index, value=value)
			cell.fill = row_fill
			cell.font = data_font
			cell.alignment = data_alignment
			cell.border = cell_border

	buffer = io.BytesIO()
	workbook.save(buffer)
	buffer.seek(0)

	response = HttpResponse(
		buffer.getvalue(),
		content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
	)
	response['Content-Disposition'] = 'attachment; filename="event_registrations.xlsx"'
	return response
